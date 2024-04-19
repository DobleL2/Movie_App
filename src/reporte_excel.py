from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
import src.analitica as analitica
import pandas as pd
from io import BytesIO

def descargar_excel(wb):
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes

def crear_reporte_excel(year):
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active

    # Añadir texto a una celda
    ws['E1'] = f"Reporte Personalizado {year}"

    # Establecer fuente, tamaño y color
    ws['E1'].font = Font(name='Amercian Typewriter',size=20,bold=True,italic=True,color='FFFFFF')

    # Ajustar el tamaño de la celda 
    ws.row_dimensions[1].height = 30
    ws.column_dimensions['E'].width = 60

    # Alinear el texto
    ws['E1'].alignment = Alignment(horizontal='center',vertical='center')

    # Aplicar un color de fondo a una celda
    ws['E1'].fill = PatternFill(start_color='050505',end_color='050505',fill_type='solid')

    img = Image(f'images/donuts_platform_year_{year}.png')
    ws.add_image(img,'A4')

    img = Image(f'images/histogram_age_year_{year}.png')
    ws.add_image(img,'F4')

    plataformas = ['Netflix','Hulu','Prime Video','Disney+']
    valores = analitica.plataforma(year)
    edades = analitica.edades_permitidas(year)
    aux = 0
    for i in range(35,39):
        ws[f'C{str(i)}'] = plataformas[aux]
        ws[f'C{str(i)}'].font = Font(name='Amercian Typewriter',size=12,bold=True,italic=True,color='050505')
        ws[f'D{str(i)}'] = valores[plataformas[aux]]
        aux +=1

    aux = 0
    for i in range(35,39):
        ws[f'L{str(i)}'] = edades.iloc[aux]['Age']
        ws[f'L{str(i)}'].font = Font(name='Amercian Typewriter',size=12,bold=True,italic=True,color='050505')
        ws[f'M{str(i)}'] = edades.iloc[aux]['Amount']
        aux +=1
        
    return wb
